import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


def read_excel(path: str) -> pd.DataFrame:
    return pd.read_excel(path, dtype=object)


def write_back_excel(df: pd.DataFrame, path: str) -> None:
    try:
        wb = load_workbook(path)
        ws = wb.active
        tables = list(getattr(ws, "_tables", {}).values()) or list(ws.tables.values())
        table_info = None
        if tables:
            t0 = tables[0]
            table_info = (t0.displayName, t0.tableStyleInfo)

        ws.delete_rows(1, ws.max_row)

        for c_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=c_idx, value=col_name)

        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        if table_info:
            name, style = table_info
            ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
            from openpyxl.worksheet.table import Table

            if name in ws.tables:
                del ws.tables[name]

            new_tbl = Table(displayName=name, ref=ref)
            if style:
                new_tbl.tableStyleInfo = style
            ws.add_table(new_tbl)

        wb.save(path)
        return
    except Exception:
        pass
    df.to_excel(path, index=False)


def iter_rows_for_sync(df: pd.DataFrame):
    if "llm_status" not in df.columns:
        for idx in df.index:
            yield idx, df.loc[idx]
        return

    status = df["llm_status"]
    mask = status.isna() | status.isnull() | status.isin(["DONE", "ERROR"])
    for idx in df[mask].index:
        yield idx, df.loc[idx]
